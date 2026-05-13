from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


OUT = Path(r"E:\develop-lcb\workspace-tools\ClipZen\docs\hwbd_standard_testcases.xlsx")


CASES = [
    {
        "id": "HWBD-001",
        "module": "仓库管理自动绑定开关",
        "name": "查询页默认值展示",
        "precondition": "进入仓库映射管理列表页，且页面数据正常加载。",
        "steps": "1. 打开查询区域。\n2. 查看两个下拉框默认值。",
        "expected": "自动绑定货权下拉框默认显示“全部”；允许吨位差下拉框默认显示“全部”。",
        "priority": "P1",
        "type": "正向",
        "note": "对应查询默认值。",
    },
    {
        "id": "HWBD-002",
        "module": "仓库管理自动绑定开关",
        "name": "查询页按启用筛选",
        "precondition": "列表页存在自动绑定货权为启用的数据。",
        "steps": "1. 自动绑定货权选择“启用”。\n2. 发起查询。",
        "expected": "列表仅展示自动绑定货权为启用的数据。",
        "priority": "P1",
        "type": "正向",
        "note": "筛选条件校验。",
    },
    {
        "id": "HWBD-003",
        "module": "仓库管理自动绑定开关",
        "name": "查询页按禁用筛选",
        "precondition": "列表页存在自动绑定货权为禁用的数据。",
        "steps": "1. 自动绑定货权选择“禁用”。\n2. 发起查询。",
        "expected": "列表仅展示自动绑定货权为禁用的数据。",
        "priority": "P1",
        "type": "正向",
        "note": "筛选条件校验。",
    },
    {
        "id": "HWBD-004",
        "module": "仓库管理自动绑定开关",
        "name": "列表展示字段",
        "precondition": "列表页存在启用和禁用两类数据。",
        "steps": "1. 打开列表页。\n2. 检查字段展示。",
        "expected": "列表展示自动绑定货权、允许吨位差字段。",
        "priority": "P1",
        "type": "正向",
        "note": "字段展示校验。",
    },
    {
        "id": "HWBD-005",
        "module": "仓库管理自动绑定开关",
        "name": "导出字段包含开关信息",
        "precondition": "列表页存在多条数据。",
        "steps": "1. 点击导出。\n2. 打开导出文件。",
        "expected": "导出列表包含自动绑定货权、允许吨位差字段。",
        "priority": "P1",
        "type": "正向",
        "note": "导出字段校验。",
    },
    {
        "id": "HWBD-006",
        "module": "仓库管理自动绑定开关",
        "name": "新增时自动绑定货权默认启用",
        "precondition": "进入新增仓库映射页面。",
        "steps": "1. 打开新增页。\n2. 查看自动绑定货权默认值。",
        "expected": "自动绑定货权默认启用，且为单选。",
        "priority": "P0",
        "type": "正向",
        "note": "新增默认值。",
    },
    {
        "id": "HWBD-007",
        "module": "仓库管理自动绑定开关",
        "name": "新增时允许吨位差默认禁用",
        "precondition": "进入新增仓库映射页面。",
        "steps": "1. 打开新增页。\n2. 查看允许吨位差默认值。",
        "expected": "允许吨位差默认禁用，且为单选。",
        "priority": "P0",
        "type": "正向",
        "note": "新增默认值。",
    },
    {
        "id": "HWBD-008",
        "module": "仓库管理自动绑定开关",
        "name": "编辑时启用切换为禁用",
        "precondition": "存在一条自动绑定货权为启用的数据。",
        "steps": "1. 打开编辑页。\n2. 将自动绑定货权改为禁用。\n3. 保存。",
        "expected": "保存成功，列表和详情显示为禁用。",
        "priority": "P0",
        "type": "正向",
        "note": "状态流转。",
    },
    {
        "id": "HWBD-009",
        "module": "仓库管理自动绑定开关",
        "name": "编辑时禁用切换为启用",
        "precondition": "存在一条自动绑定货权为禁用的数据。",
        "steps": "1. 打开编辑页。\n2. 将自动绑定货权改为启用。\n3. 保存。",
        "expected": "保存成功，列表和详情显示为启用。",
        "priority": "P0",
        "type": "正向",
        "note": "状态流转。",
    },
    {
        "id": "HWBD-010",
        "module": "自动绑定判定逻辑",
        "name": "找钢网-预售预购租户不参与自动绑定",
        "precondition": "租户配置为找钢网-预售预购租户。",
        "steps": "1. 触发自动绑定任务。\n2. 观察任务结果。",
        "expected": "该租户的数据不参与货权自动绑定。",
        "priority": "P0",
        "type": "正向",
        "note": "租户级排除规则。",
    },
    {
        "id": "HWBD-011",
        "module": "自动绑定判定逻辑",
        "name": "未开启自动绑定时仅执行收货单完结",
        "precondition": "租户非找钢网-预售预购，且开关关闭。",
        "steps": "1. 触发收货单明细定时任务。\n2. 查看收货单状态。",
        "expected": "收货单按完结逻辑处理，不进入货权自动绑定。",
        "priority": "P0",
        "type": "正向",
        "note": "开关关闭路径。",
    },
    {
        "id": "HWBD-012",
        "module": "自动绑定判定逻辑",
        "name": "开启自动绑定但售后状态为0实提时不自动绑定",
        "precondition": "租户非找钢网-预售预购，开关开启，售后状态=0实提。",
        "steps": "1. 触发任务。\n2. 查看结果。",
        "expected": "数据不参与自动绑定，进入手动处理路径。",
        "priority": "P0",
        "type": "反向",
        "note": "售后状态拦截。",
    },
    {
        "id": "HWBD-013",
        "module": "自动绑定判定逻辑",
        "name": "订单标签命中时不自动绑定",
        "precondition": "租户非找钢网-预售预购，开关开启，售后状态!=0实提。",
        "steps": "1. 让过户单订单标签命中预售订单/运营敏采转销售/运营敏采（系统）。\n2. 触发任务。",
        "expected": "数据不参与自动绑定，进入手动处理路径。",
        "priority": "P0",
        "type": "反向",
        "note": "订单标签拦截。",
    },
    {
        "id": "HWBD-014",
        "module": "自动绑定判定逻辑",
        "name": "仓储确认等前置条件不满足时不自动绑定",
        "precondition": "租户非找钢网-预售预购，开关开启。",
        "steps": "1. 构造仓储确认状态、关联状态、过户状态、售后状态不满足条件的数据。\n2. 触发任务。",
        "expected": "数据不参与自动绑定，进入手动处理路径。",
        "priority": "P0",
        "type": "反向",
        "note": "前置条件拦截。",
    },
    {
        "id": "HWBD-015",
        "module": "自动绑定判定逻辑",
        "name": "过户单明细已关联吨位或数量不为0时不自动绑定",
        "precondition": "满足前置条件，且过户单明细已关联吨位/数量!=0。",
        "steps": "1. 触发任务。\n2. 查看结果。",
        "expected": "数据不参与自动绑定，进入手动处理路径。",
        "priority": "P0",
        "type": "反向",
        "note": "明细已产生业务操作。",
    },
    {
        "id": "HWBD-016",
        "module": "自动绑定判定逻辑",
        "name": "收货单条件不满足时不自动绑定",
        "precondition": "过户单明细关联为0，且收货单状态不满足批准未完结等条件。",
        "steps": "1. 触发任务。\n2. 查看结果。",
        "expected": "数据不参与自动绑定，进入手动处理路径。",
        "priority": "P0",
        "type": "反向",
        "note": "收货单约束拦截。",
    },
    {
        "id": "HWBD-017",
        "module": "自动绑定判定逻辑",
        "name": "过户单与收货单日期超出1天时不自动绑定",
        "precondition": "满足前置条件。",
        "steps": "1. 构造过户单进入云仓日期与收货单收货日期差值大于1天的数据。\n2. 触发任务。",
        "expected": "数据不参与自动绑定，进入手动处理路径。",
        "priority": "P0",
        "type": "边界",
        "note": "日期边界。",
    },
    {
        "id": "HWBD-018",
        "module": "自动绑定判定逻辑",
        "name": "供应商与原货主不一致时不自动绑定",
        "precondition": "满足前置条件。",
        "steps": "1. 构造供应商名称与原货主归一化后仍不一致的数据。\n2. 触发任务。",
        "expected": "数据不参与自动绑定，进入手动处理路径。",
        "priority": "P0",
        "type": "反向",
        "note": "供应商归一化校验。",
    },
    {
        "id": "HWBD-019",
        "module": "自动绑定判定逻辑",
        "name": "SKU不匹配时不自动绑定",
        "precondition": "满足前置条件。",
        "steps": "1. 构造过户单明细与收货单明细 SKU 不匹配的数据。\n2. 触发任务。",
        "expected": "数据不参与自动绑定，进入手动处理路径。",
        "priority": "P0",
        "type": "反向",
        "note": "SKU匹配校验。",
    },
    {
        "id": "HWBD-020",
        "module": "自动绑定判定逻辑",
        "name": "热卷或冷镀品线捆包号校验",
        "precondition": "满足前置条件，品线为热卷或冷镀。",
        "steps": "1. 构造收货单捆包号为123_456、过户单为123的数据。\n2. 触发任务。\n3. 再构造过户单为12的数据并触发任务。",
        "expected": "123与123_456可匹配；12与123_456不匹配。",
        "priority": "P0",
        "type": "边界",
        "note": "捆包号规则。",
    },
    {
        "id": "HWBD-021",
        "module": "自动绑定判定逻辑",
        "name": "非热卷或冷镀无需检验捆包号",
        "precondition": "满足前置条件，品线非热卷或冷镀。",
        "steps": "1. 构造捆包号不一致但其他条件满足的数据。\n2. 触发任务。",
        "expected": "无需检验捆包号，按其他规则继续判定。",
        "priority": "P1",
        "type": "正向",
        "note": "品线差异化规则。",
    },
    {
        "id": "HWBD-022",
        "module": "自动绑定判定逻辑",
        "name": "允许吨位差开启时满足容错条件自动绑定",
        "precondition": "允许吨位差=启用，且数量一致、吨位差在阈值内、双方已关联吨位和件数为0。",
        "steps": "1. 构造符合容错条件的数据。\n2. 触发任务。",
        "expected": "过户单明细与收货单明细自动绑定并记录日志。",
        "priority": "P0",
        "type": "正向",
        "note": "吨位差启用路径。",
    },
    {
        "id": "HWBD-023",
        "module": "自动绑定判定逻辑",
        "name": "允许吨位差开启时超出阈值不绑定",
        "precondition": "允许吨位差=启用。",
        "steps": "1. 构造吨位差超出阈值的数据。\n2. 触发任务。",
        "expected": "数据不参与自动绑定，进入手动处理路径。",
        "priority": "P0",
        "type": "反向",
        "note": "阈值边界。",
    },
    {
        "id": "HWBD-024",
        "module": "自动绑定判定逻辑",
        "name": "允许吨位差关闭时必须严格一致",
        "precondition": "允许吨位差=禁用。",
        "steps": "1. 构造吨位存在差异的数据。\n2. 触发任务。",
        "expected": "差异数据不自动绑定；只有严格一致的数据才会绑定。",
        "priority": "P0",
        "type": "反向",
        "note": "严格一致规则。",
    },
    {
        "id": "HWBD-025",
        "module": "仓储确认定时任务",
        "name": "满足所有条件时自动确认",
        "precondition": "过户单满足待确认、部分关联、关联件数等于预提件数且吨位差在阈值内。",
        "steps": "1. 触发仓储确认定时任务。\n2. 查看状态与日志。",
        "expected": "仓储确认状态变为已确认，并记录操作日志。",
        "priority": "P0",
        "type": "正向",
        "note": "定时任务确认路径。",
    },
    {
        "id": "HWBD-026",
        "module": "仓储确认定时任务",
        "name": "仅满足部分条件时保持不变",
        "precondition": "过户单仅满足部分条件。",
        "steps": "1. 触发仓储确认定时任务。\n2. 查看状态。",
        "expected": "仓储确认状态保持不变。",
        "priority": "P1",
        "type": "反向",
        "note": "部分命中不应误确认。",
    },
    {
        "id": "HWBD-027",
        "module": "收货单明细定时任务",
        "name": "满足所有条件时自动完结",
        "precondition": "收货单满足部分关联、未关联件数为0、吨位差在阈值内。",
        "steps": "1. 触发收货单明细定时任务。\n2. 查看状态与日志。",
        "expected": "收货单关联状态保持不变，明细完结状态变为已完结，并记录系统日志。",
        "priority": "P0",
        "type": "正向",
        "note": "定时任务完结路径。",
    },
    {
        "id": "HWBD-028",
        "module": "收货单明细定时任务",
        "name": "仅满足部分条件时保持不变",
        "precondition": "收货单仅满足部分条件。",
        "steps": "1. 触发收货单明细定时任务。\n2. 查看状态。",
        "expected": "执行后状态保持不变。",
        "priority": "P1",
        "type": "反向",
        "note": "部分命中不应误完结。",
    },
    {
        "id": "HWBD-029",
        "module": "历史数据初始化以及检查",
        "name": "自动绑定开关历史数据初始化验证",
        "precondition": "存在历史未开启数据。",
        "steps": "1. 将历史数据初始化为开启。\n2. 触发任务。\n3. 查看结果。",
        "expected": "历史数据初始化后任务执行正常，无异常数据漏绑或误绑。",
        "priority": "P1",
        "type": "回归",
        "note": "历史数据迁移。",
    },
    {
        "id": "HWBD-030",
        "module": "历史数据初始化以及检查",
        "name": "允许吨位差历史数据初始化验证",
        "precondition": "存在历史禁用数据。",
        "steps": "1. 将历史数据初始化为启用。\n2. 触发任务。\n3. 查看结果。",
        "expected": "历史数据初始化后任务执行正常。",
        "priority": "P1",
        "type": "回归",
        "note": "历史数据迁移。",
    },
    {
        "id": "HWBD-031",
        "module": "手动相关功能回归",
        "name": "手动货权绑定查询关联量",
        "precondition": "进入手动货权绑定页面。",
        "steps": "1. 执行查询关联。\n2. 查看过户单和收货单关联量。",
        "expected": "查询结果正确展示关联量检查结果。",
        "priority": "P1",
        "type": "正向",
        "note": "手动绑定查询。",
    },
    {
        "id": "HWBD-032",
        "module": "手动相关功能回归",
        "name": "手动仓储确认和驳回检查状态",
        "precondition": "存在可操作的过户单。",
        "steps": "1. 执行手动仓储确认。\n2. 再执行手动仓储驳回。\n3. 查看状态。",
        "expected": "状态校验正确，操作结果符合预期。",
        "priority": "P1",
        "type": "正向",
        "note": "手动状态流转。",
    },
    {
        "id": "HWBD-033",
        "module": "手动相关功能回归",
        "name": "解绑后不再参与自动绑定",
        "precondition": "存在已执行过解绑的明细。",
        "steps": "1. 对明细执行解绑。\n2. 触发自动绑定任务。\n3. 查看结果。",
        "expected": "该明细不再参与自动绑定执行任务。",
        "priority": "P0",
        "type": "回归",
        "note": "解绑排除规则。",
    },
    {
        "id": "HWBD-034",
        "module": "手动相关功能回归",
        "name": "手动完结后状态同步",
        "precondition": "存在可完结收货单。",
        "steps": "1. 执行收货单完结。\n2. 查看关联查询和明细状态。",
        "expected": "关联查询无该数据，收货单明细完结状态同步成功。",
        "priority": "P0",
        "type": "正向",
        "note": "手动完结后同步。",
    },
    {
        "id": "HWBD-035",
        "module": "切换",
        "name": "RPA逐步关停切换",
        "precondition": "线上存在 RPA 相关流程。",
        "steps": "1. 逐个关闭 RPA。\n2. 单个流程验证无误后继续。\n3. 最终全量切换。",
        "expected": "切换过程无异常，全部流程验证通过后再全量切换。",
        "priority": "P0",
        "type": "变更",
        "note": "上线切换。",
    },
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    headers = [
        "用例ID",
        "模块",
        "用例名称",
        "前置条件",
        "测试步骤",
        "预期结果",
        "优先级",
        "类型",
        "备注",
    ]
    ws.append(headers)
    for case in CASES:
        ws.append(
            [
                case["id"],
                case["module"],
                case["name"],
                case["precondition"],
                case["steps"],
                case["expected"],
                case["priority"],
                case["type"],
                case["note"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {
        "A": 12,
        "B": 24,
        "C": 32,
        "D": 30,
        "E": 40,
        "F": 38,
        "G": 10,
        "H": 10,
        "I": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    notes = wb.create_sheet("说明")
    notes.append(["说明项", "内容"])
    notes.append(["来源", "由 xmind 内容整理为标准测试用例表"])
    notes.append(["特点", "按测试执行需要补齐了前置条件、步骤、预期结果、优先级和类型"])
    notes.append(["建议", "如果你要，我可以继续给这版补“测试数据”和“执行人/执行环境”两列"])
    for cell in notes[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in notes.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    notes.column_dimensions["A"].width = 14
    notes.column_dimensions["B"].width = 90
    notes.freeze_panes = "A2"

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
