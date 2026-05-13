from __future__ import annotations

import json
import os
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


XLSX_OUT = Path(r"E:\develop-lcb\workspace-tools\ClipZen\docs\hwbd_testcases_final.xlsx")


def load_root_topic(xmind_path: str) -> dict:
    with zipfile.ZipFile(xmind_path) as zf:
        return json.loads(zf.read("content.json").decode("utf-8"))[0]["rootTopic"]


def collect_nodes(root: dict) -> list[dict]:
    rows: list[dict] = []

    def walk(node: dict, path: list[str] | None = None, depth: int = 0) -> None:
        path = (path or []) + [node.get("title", "")]
        children = node.get("children", {}).get("attached", []) + node.get("children", {}).get(
            "detached", []
        )
        rows.append(
            {
                "depth": depth,
                "title": node.get("title", ""),
                "path": " / ".join(path),
                "is_leaf": "是" if not children else "否",
            }
        )
        for child in children:
            walk(child, path, depth + 1)

    walk(root)
    return rows


def build_workbook(rows: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Outline"

    headers = [
        "\u5c42\u7ea7",
        "\u8282\u70b9\u6807\u9898",
        "\u5b8c\u6574\u8def\u5f84",
        "\u53f6\u5b50\u8282\u70b9",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row["depth"], row["title"], row["path"], row["is_leaf"]])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 110
    ws.column_dimensions["D"].width = 10
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    flat = wb.create_sheet("Leafs")
    flat_headers = [
        "\u7528\u4f8bID",
        "\u4e00\u7ea7\u6a21\u5757",
        "\u4e8c\u7ea7\u6a21\u5757",
        "\u4e09\u7ea7\u6a21\u5757",
        "\u56db\u7ea7\u6a21\u5757",
        "\u8282\u70b9\u6807\u9898",
        "\u5b8c\u6574\u8def\u5f84",
    ]
    flat.append(flat_headers)
    idx = 1
    for row in rows:
        if row["is_leaf"] != "是":
            continue
        parts = [p for p in row["path"].split(" / ") if p]
        path_cols = parts[1:5] + [""] * 4
        flat.append(
            [
                f"HWBD-{idx:03d}",
                path_cols[0],
                path_cols[1],
                path_cols[2],
                path_cols[3],
                row["title"],
                row["path"],
            ]
        )
        idx += 1

    for cell in flat[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in flat.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {"A": 12, "B": 22, "C": 22, "D": 22, "E": 22, "F": 42, "G": 110}
    for col, width in widths.items():
        flat.column_dimensions[col].width = width
    flat.freeze_panes = "A2"
    flat.auto_filter.ref = flat.dimensions

    notes = wb.create_sheet("Notes")
    notes.append(["\u8bf4\u660e\u9879", "\u5185\u5bb9"])
    notes.append(["\u6765\u6e90", "\u76f4\u63a5\u4ece xmind \u539f\u6587\u4ef6\u8bfb\u53d6\uff0c\u907f\u514d\u4e2d\u6587\u7f16\u7801\u635f\u574f"])
    notes.append(["\u7528\u9014", "\u5148\u4fdd\u8bc1\u4e2d\u6587\u53ef\u89c1\uff0c\u518d\u6309\u9700\u8981\u7ee7\u7eed\u6574\u7406\u6210\u5b8c\u6574\u6d4b\u8bd5\u7528\u4f8b\u8868"])
    notes.append(["\u5efa\u8bae", "\u5982\u679c\u4f60\u8981\u6b63\u5f0f\u6d4b\u8bd5\u6267\u884c\u7248\uff0c\u6211\u53ef\u4ee5\u7ee7\u7eed\u628a\u6bcf\u4e2a\u53f6\u5b50\u8282\u70b9\u6269\u6210\u201c\u524d\u7f6e\u6761\u4ef6/\u6b65\u9aa4/\u9884\u671f\u7ed3\u679c\u201d"])

    for cell in notes[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in notes.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    notes.column_dimensions["A"].width = 16
    notes.column_dimensions["B"].width = 90
    notes.freeze_panes = "A2"

    return wb


def main() -> None:
    xmind_path = os.environ.get("XMIN_PATH")
    if not xmind_path:
        raise SystemExit("XMIN_PATH is required")
    root = load_root_topic(xmind_path)
    rows = collect_nodes(root)
    wb = build_workbook(rows)
    wb.save(XLSX_OUT)
    print(XLSX_OUT)


if __name__ == "__main__":
    main()
